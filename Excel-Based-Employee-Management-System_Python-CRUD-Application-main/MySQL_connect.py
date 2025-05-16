# https://myjamong.tistory.com/53#google_vignette

import pymysql
from openpyxl import Workbook
from openpyxl import load_workbook

# create table test(num int(11), name varchar(10));

class Test:
    def __init__(self, num, name):
        self.num = num
        self.name = name

#전체 Select
def select_all():
    conn = pymysql.connect(host='localhost', user='root', password='dark##2993', db='python', charset='utf8')
    try:
        with conn.cursor() as curs:
            sql = "select * from test"
            curs.execute(sql)
            rs = curs.fetchall()
            for row in rs:
                print(row)
    finally:
        conn.close()

# DB Insert
def insert_test(test_obj):
    conn = pymysql.connect(host='localhost', user='root', password='dark##2993', db='python', charset='utf8')
    try:
        with conn.cursor() as curs:
            sql = 'insert into test values(%s, %s)'
            curs.execute(sql, (test_obj.num, test_obj.name))
        conn.commit()
    finally :
        conn.close()
    

# DB Delete
def delete_test(num):
    conn = pymysql.connect(host='localhost', user='root', password='dark##2993', db='python', charset='utf8')
    try:
        with conn.cursor() as curs:
            sql = 'delete from test where num=%s'
            curs.execute(sql, num)
        conn.commit()
    finally:
        conn.close()

# DB Delete All
def delete_all():
    conn = pymysql.connect(host='localhost', user='root', password='dark##2993', db='python', charset='utf8')
    try:
        with conn.cursor() as curs:
            sql = 'delete from test'
            curs.execute(sql)
        conn.commit()
    finally:
        conn.close()

# DB Update
def update_test(test_obj):
    conn = pymysql.connect(host='localhost', user='root', password='dark##2993', db='python', charset='utf8')
    try:
        with conn.cursor() as curs:
            sql = 'update test set name=%s where num=%s'
            curs.execute(sql, (test_obj.name, test_obj.num))
        conn.commit()
    finally:
        conn.close()

# 전체 select 하여 엑셀 파일 쓰기
def select_all_to_excel():
    conn = pymysql.connect(host='localhost', user='root', password='dark##2993', db='python', charset='utf8')
    try:
        with conn.cursor() as curs:
            sql = "select * from test"
            curs.execute(sql)
            rs = curs.fetchall()

            wb = Workbook()
            ws = wb.active

            #첫행 입력
            ws.append(('번호','이름'))

            #DB 모든 데이터 엑셀로
            for row in rs:
                ws.append(row)

            wb.save('P:/임시/숫자.xlsx')
    finally:
        conn.close()
        wb.close()

#엑셀파일 DB Insert
def insert_excel_to_db():
    conn = pymysql.connect(host='localhost', user='root', password='dark##2993', db='python', charset='utf8')
    try:
        with conn.cursor() as curs:
            sql = 'insert into test values(%s, %s)'

            wb = load_workbook('P:/임시/숫자.xlsx',data_only=True)
            ws = wb['Sheet']

            iter_rows = iter(ws.rows)
            next(iter_rows)
            for row in iter_rows:
                curs.execute(sql, (row[0].value, row[1].value))
            conn.commit()
    finally:
        conn.close()
        wb.close()


if __name__ == "__main__":
    # for i in range(1, 1000):
    #     test = Test(i, str(i) + '이름')
    #     insert_test(test)
    
    # DB → 엑셀파일
    # select_all_to_excel()

    # select_all()
    # delete_test(999)
    
    # test = Test(998, '___이름')
    # update_test(test)

    # delete_all()

    insert_excel_to_db()
    select_all()